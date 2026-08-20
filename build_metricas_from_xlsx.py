import json
from collections import Counter, defaultdict
from datetime import date, datetime, timedelta
from pathlib import Path

from openpyxl import load_workbook

SOURCE = Path('/home/ubuntu/upload/Ouvidorias-13082026082544.xlsx')
TARGET = Path('/home/ubuntu/falabr-facil/client/public/metricas.json')
START = date(2021, 8, 1)
CATEGORY_ORDER = ['Reclamação', 'Solicitação', 'Denúncia', 'Informação', 'Elogio', 'Sugestão', 'Doação', 'Simplifique']
FINAL = {'Concluído', 'Cancelado'}

def clean(value):
    return value.strip() if isinstance(value, str) else value

def parse_date(value):
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        for fmt in ('%d/%m/%Y %H:%M:%S', '%d/%m/%Y'):
            try:
                return datetime.strptime(value, fmt).date()
            except ValueError:
                continue
    return None

workbook = load_workbook(SOURCE, read_only=True, data_only=True)
sheet = workbook['Dados']
headers = [cell.value for cell in next(sheet.iter_rows(values_only=False))]
index = {name: position for position, name in enumerate(headers)}

daily = defaultdict(lambda: {'f': 0, 'fab': 0, 'fcc': 0, 'fcn': 0, 'cat': Counter()})
monthly_secretariat = defaultdict(lambda: defaultdict(lambda: {'reg': 0, 'ab': 0, 'cc': 0}))
monthly_subject = defaultdict(Counter)
secretariat_totals = Counter()
subject_totals = Counter()
quality = Counter()
latest = None

for row in sheet.iter_rows(min_row=2, values_only=True):
    received = parse_date(row[index['Recebido em']])
    if not received or received < START:
        if received and received.year < 2000:
            quality['datas_recebimento_invalidas'] += 1
        continue
    status = clean(row[index['Status']]) or '(vazio)'
    secretariat = clean(row[index['Secretaria']]) or 'Não informado'
    subject = clean(row[index['Assunto']]) or 'Não informado'
    category = clean(row[index['Categoria']]) or 'Não informado'
    finalized = parse_date(row[index['Finalizado em']])
    opinion = clean(row[index['Opinião']]) or '(vazio)'
    key = received.isoformat()
    month = key[:7]
    latest = max(latest, received) if latest else received
    metrics = daily[key]
    metrics['f'] += 1
    metrics['cat'][category] += 1
    monthly_secretariat[month][secretariat]['reg'] += 1
    monthly_subject[month][subject] += 1
    secretariat_totals[secretariat] += 1
    subject_totals[subject] += 1
    if status == 'Concluído':
        metrics['fcc'] += 1
        monthly_secretariat[month][secretariat]['cc'] += 1
        if not finalized:
            quality['concluidos_sem_data_finalizacao'] += 1
    elif status == 'Cancelado':
        metrics['fcn'] += 1
    else:
        metrics['fab'] += 1
    if not finalized:
        quality['sem_data_finalizacao'] += 1
    if opinion in ('Positivo', 'Negativo'):
        quality[f'opiniao_{opinion.lower()}'] += 1
    elif opinion in ('Sem opinião', '(vazio)'):
        quality['sem_avaliacao'] += 1

categories = [category for category in CATEGORY_ORDER if any(category in row['cat'] for row in daily.values())]
secretariats = [item for item, _ in secretariat_totals.most_common()]
subjects = [item for item, _ in subject_totals.most_common(20)]
secretariat_index = {name: pos for pos, name in enumerate(secretariats)}
subject_index = {name: pos for pos, name in enumerate(subjects)}

serialized_daily = {}
for day, values in sorted(daily.items()):
    serialized_daily[day] = {
        'f': values['f'], 'fab': values['fab'], 'fcc': values['fcc'], 'fcn': values['fcn'],
        'cat': [values['cat'][category] for category in categories]
    }

serialized_monthly = {}
for month in sorted(monthly_secretariat):
    sec = monthly_secretariat[month]
    ass = monthly_subject[month]
    serialized_monthly[month] = {
        'sec': {
            'reg': [sec[name]['reg'] for name in secretariats],
            'ab': [sec[name]['ab'] for name in secretariats],
            'cc': [sec[name]['cc'] for name in secretariats],
        },
        'ass': {'reg': [ass[name] for name in subjects]},
    }

open_over_30 = 0
if latest:
    cutoff = latest - timedelta(days=30)
    for day, values in daily.items():
        if date.fromisoformat(day) <= cutoff:
            open_over_30 += values['fab']

payload = {
    'meta': {
        'gerado_em': latest.isoformat() if latest else None,
        'janela_inicio': START.isoformat(),
        'janela_fim': latest.isoformat() if latest else None,
        'categorias': categories,
        'secretarias': secretariats,
        'assuntos': subjects,
        'fonte': 'Planilha Ouvidorias-13082026082544.xlsx — agregada e anonimizada',
    },
    'diario': serialized_daily,
    'mensal_dim': serialized_monthly,
    'abertos_mais_30d': {'referencia': latest.isoformat() if latest else None, 'total': open_over_30},
    'qualidade': dict(quality),
}
TARGET.write_text(json.dumps(payload, ensure_ascii=False, separators=(',', ':')))
print(json.dumps({'target': str(TARGET), 'days': len(serialized_daily), 'records': sum(item['f'] for item in serialized_daily.values()), 'open_over_30': open_over_30, 'quality': dict(quality)}, ensure_ascii=False))
