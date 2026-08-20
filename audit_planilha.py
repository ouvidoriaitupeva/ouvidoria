from collections import Counter
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

path = Path('/home/ubuntu/upload/Ouvidorias-13082026082544.xlsx')
workbook = load_workbook(path, read_only=True, data_only=True)
sheet = workbook['Dados']
headers = [cell.value for cell in next(sheet.iter_rows(values_only=False))]
index = {name: position for position, name in enumerate(headers)}

required = ['Recebido em', 'Status', 'Secretaria', 'Assunto', 'Categoria', 'Finalizado em', 'Opinião', 'Visível']
print('headers_found', {name: name in index for name in required})

def value(row, name):
    return row[index[name]] if name in index else None

def normalize_date(raw):
    if isinstance(raw, datetime):
        return raw.date()
    if isinstance(raw, str):
        for fmt in ('%d/%m/%Y %H:%M:%S', '%d/%m/%Y'):
            try:
                return datetime.strptime(raw, fmt).date()
            except ValueError:
                pass
    return None

summary = Counter()
status = Counter()
category = Counter()
secretariat = Counter()
opinion = Counter()
visible = Counter()
rows_since_json_start = Counter()
last_12 = Counter()
last_12_opinion = Counter()
date_min = None
date_max = None
finalized_dates = 0
blanks = Counter()
received_year = Counter()
finalized_by_status = Counter()
missing_finalized_by_status = Counter()
invalid_received_dates = 0

for row in sheet.iter_rows(min_row=2, values_only=True):
    summary['rows'] += 1
    received = normalize_date(value(row, 'Recebido em'))
    finalized = normalize_date(value(row, 'Finalizado em'))
    if received:
        date_min = received if date_min is None or received < date_min else date_min
        date_max = received if date_max is None or received > date_max else date_max
        received_year[received.year] += 1
        if received.year < 2000:
            invalid_received_dates += 1
    else:
        summary['blank_received'] += 1
    if finalized:
        finalized_dates += 1
    stat = (value(row, 'Status') or '').strip() if isinstance(value(row, 'Status'), str) else value(row, 'Status') or '(vazio)'
    cat = (value(row, 'Categoria') or '').strip() if isinstance(value(row, 'Categoria'), str) else value(row, 'Categoria') or '(vazio)'
    sec = (value(row, 'Secretaria') or '').strip() if isinstance(value(row, 'Secretaria'), str) else value(row, 'Secretaria') or '(vazio)'
    status[stat] += 1
    category[cat] += 1
    secretariat[sec] += 1
    opinion[value(row, 'Opinião') or '(vazio)'] += 1
    visible[value(row, 'Visível') or '(vazio)'] += 1
    for field in ['Status', 'Secretaria', 'Assunto', 'Categoria', 'Finalizado em', 'Opinião']:
        raw = value(row, field)
        if raw is None or (isinstance(raw, str) and not raw.strip()):
            blanks[field] += 1
    if finalized:
        finalized_by_status[stat] += 1
    else:
        missing_finalized_by_status[stat] += 1
    if received and received >= datetime(2021, 8, 1).date():
        rows_since_json_start['all'] += 1
        rows_since_json_start[f'status:{stat}'] += 1
        rows_since_json_start[f'category:{cat}'] += 1
    if received and datetime(2025, 9, 1).date() <= received <= datetime(2026, 8, 13).date():
        last_12['all'] += 1
        last_12[f'status:{stat}'] += 1
        last_12[f'category:{cat}'] += 1
        last_12_opinion[value(row, 'Opinião') or '(vazio)'] += 1

print('date_range', date_min, date_max)
print('total_rows', summary['rows'], 'blank_received', summary['blank_received'], 'with_finalized_date', finalized_dates)
print('status_top', status.most_common(20))
print('category_top', category.most_common(20))
print('opinion', opinion.most_common(10))
print('visible', visible.most_common(10))
print('blank_fields', dict(blanks))
print('received_years', sorted(received_year.items()))
print('invalid_received_dates_before_2000', invalid_received_dates)
print('finalized_by_status', dict(finalized_by_status))
print('missing_finalized_by_status', dict(missing_finalized_by_status))
print('json_window_all', rows_since_json_start['all'])
print('json_window_status', [(key.replace('status:', ''), value) for key, value in rows_since_json_start.items() if key.startswith('status:')])
print('last12_all', last_12['all'])
print('last12_status', [(key.replace('status:', ''), value) for key, value in last_12.items() if key.startswith('status:')])
print('last12_categories', sorted([(key.replace('category:', ''), value) for key, value in last_12.items() if key.startswith('category:')], key=lambda item: -item[1]))
print('last12_opinion', last_12_opinion.most_common(10))
print('top_secretariats_last12_not_calculated_in_first_pass')
